from __future__ import annotations

from openpyxl import load_workbook

from app.services.admin_service import AdminService
from app.services.export_service import CARD_COLUMNS, save_cards_xlsx
from tests.helpers import import_restricted_seed, import_seed, migrated_db


def test_cards_export_is_russian_editable_and_round_trips(tmp_path):
    db = migrated_db(tmp_path)
    import_seed(db)
    import_restricted_seed(db)
    service = AdminService(db)
    path = tmp_path / "cards_export.xlsx"

    save_cards_xlsx(service.export_rows(), path, service.export_items())

    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Флирт",
        "Разогрев",
        "Секс",
        "BDSM",
        "Экстрим",
        "Реквизит",
        "Справочники",
    ]
    assert [cell.value for cell in workbook["Флирт"][1]] == CARD_COLUMNS
    assert workbook["Флирт"].freeze_panes == "A2"
    assert workbook["Флирт"].data_validations.count >= 8
    assert workbook["Экстрим"].max_row == 23
    assert workbook["Реквизит"]["B2"].value
    assert "check-in" not in " ".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )

    report = service.import_content(111, str(path), dry_run=True)
    assert report.added_or_updated == 218
    assert report.items_added_or_updated == 16
    assert report.warnings_count == 0
    db.close()

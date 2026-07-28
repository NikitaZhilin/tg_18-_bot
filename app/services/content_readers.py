from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from typing import Any

from app.domain import parse_json_list
from app.labels import (
    BODY_LOAD_NAMES,
    CATEGORY_NAMES,
    INTENSITY_NAMES,
    ITEM_MODE_NAMES,
    POSE_DIFFICULTY_NAMES,
    POSE_FAMILY_NAMES,
    RISK_TAG_NAMES,
    SPACE_NAMES,
    code_from_label,
)


REQUIRED_COLUMNS = {"external_id", "level", "category", "text"}
RISK_KEYWORDS = {
    "breath_control": ("дыхани", "удуш", "перекры"),
    "choking": ("горло", "шею", "шея"),
    "unsafe_wax": ("воск", "свеч"),
    "no_quick_release_restraint": ("свяж", "связ", "верев"),
    "unbounded_humiliation": ("униж",),
    "intoxication": ("алког", "опьян"),
    "recording_without_consent": ("фото", "видео", "запиш"),
    "public_nonconsenting_people": ("публич", "посторон"),
    "injury": ("боль", "удар", "след", "синяк"),
}
ITEM_ALIASES = {
    "лед": "ice",
    "масло": "oil",
    "веревка": "rope",
    "верёвка": "rope",
    "повязка": "blindfold",
    "вибратор": "vibrator",
    "наушники": "headphones",
    "свеча": "candle",
    "свечи": "candle",
    "зажимы": "clamps",
    "еда": "food",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or not REQUIRED_COLUMNS.issubset(set(reader.fieldnames)):
            missing = ", ".join(sorted(REQUIRED_COLUMNS - set(reader.fieldnames or [])))
            raise ValueError(f"missing required columns: {missing}")
        return [dict(row) for row in reader]


def read_xlsx(path: Path) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.active and _english_headers(workbook.active):
            rows = list(workbook.active.iter_rows(values_only=True))
            headers = [str(cell or "").strip() for cell in rows[0]]
            return (
                [
                    {
                        headers[index]: (
                            "" if index >= len(raw) or raw[index] is None else str(raw[index])
                        )
                        for index in range(len(headers))
                    }
                    for raw in rows[1:]
                ],
                [],
            )

        item_rows = (
            _read_russian_item_sheet(workbook["Реквизит"])
            if "Реквизит" in workbook.sheetnames
            else []
        )
        item_codes = {
            str(item["name"]).strip().casefold(): str(item["code"])
            for item in item_rows
        }
        result: list[dict[str, str]] = []
        sheet_levels = {
            "Флирт": 1,
            "Разогрев": 2,
            "Секс": 3,
            "BDSM": 4,
            "Экстрим": 4,
            "Закрытые темы": 4,
        }
        for sheet_name, level in sheet_levels.items():
            if sheet_name not in workbook.sheetnames:
                continue
            result.extend(
                _read_russian_card_sheet(
                    workbook[sheet_name],
                    level=level,
                    extreme=sheet_name in {"Экстрим", "Закрытые темы"},
                    item_codes=item_codes,
                )
            )
        if not result:
            raise ValueError(
                "В XLSX не найдены листы «Флирт», «Разогрев», «Секс», «BDSM» или «Экстрим»."
            )
        return result, item_rows
    finally:
        workbook.close()


def read_docx(path: Path) -> list[dict[str, str]]:
    from docx import Document

    document = Document(path)
    rows: list[dict[str, str]] = []
    level = 1
    counter = 1
    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        lower = text.lower()
        detected_level = next(
            (candidate for candidate in range(1, 5) if f"уровень {candidate}" in lower),
            None,
        )
        if detected_level:
            level = detected_level
            continue
        if len(text) < 12:
            continue
        rows.append(
            {
                "external_id": f"docx_l{level}_{counter:03d}",
                "level": str(level),
                "category": "task",
                "intensity": "light" if level < 3 else "medium",
                "title": "",
                "text": _strip_docx_metadata(text),
                "required_items": ",".join(_extract_items(text)),
                "timer_seconds": str(_extract_timer_seconds(text) or ""),
                "risk_tags": ",".join(_detect_risk_tags(text)),
                "collection": "base_tasks",
                "review_status": "needs_review",
                "is_enabled": "0",
            }
        )
        counter += 1
    return rows


def _extract_items(text: str) -> list[str]:
    match = re.search(r"\[?\s*Реквизит\s*:\s*([^\]\n]+)\]?", text, flags=re.IGNORECASE)
    if not match:
        return []
    raw_value = re.split(r"\bТаймер\s*:", match.group(1), flags=re.IGNORECASE)[0]
    return [
        ITEM_ALIASES.get(raw.lower().strip(), raw.lower().strip())
        for raw in parse_json_list(raw_value)
    ]


def _extract_timer_seconds(text: str) -> int | None:
    match = re.search(
        r"\[?\s*Таймер\s*:\s*(\d+)\s*(сек|секунд|мин|минут)?",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    value = int(match.group(1))
    return value * 60 if (match.group(2) or "сек").lower().startswith("мин") else value


def _detect_risk_tags(text: str) -> list[str]:
    lower = text.lower()
    return [
        tag
        for tag, keywords in RISK_KEYWORDS.items()
        if any(keyword in lower for keyword in keywords)
    ]


def _strip_docx_metadata(text: str) -> str:
    text = re.sub(r"\[?\s*Реквизит\s*:\s*[^\]\n]+\]?", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\[?\s*Таймер\s*:\s*[^\]\n]+\]?", "", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def _english_headers(sheet) -> bool:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return REQUIRED_COLUMNS.issubset({str(value or "").strip() for value in first_row})


def _sheet_dict_rows(sheet) -> list[tuple[int, dict[str, object]]]:
    raw_rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(raw_rows, ())]
    result = []
    for row_number, raw in enumerate(raw_rows, start=2):
        values = {
            header: raw[index] if index < len(raw) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in values.values()):
            result.append((row_number, values))
    return result


def _read_russian_card_sheet(
    sheet,
    *,
    level: int,
    extreme: bool,
    item_codes: dict[str, str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_number, values in _sheet_dict_rows(sheet):
        text = str(values.get("Текст карточки") or "").strip()
        if not text:
            continue
        external_id = str(values.get("Код карточки") or "").strip()
        if not external_id:
            digest = hashlib.sha1(
                f"{sheet.title}|{row_number}|{text}".encode("utf-8")
            ).hexdigest()[:12]
            external_id = f"xlsx_{digest}"
        category = _mapped_code(values.get("Тип карточки"), CATEGORY_NAMES, "task")
        intensity = _mapped_code(
            values.get("Интенсивность"),
            INTENSITY_NAMES,
            "hard" if extreme else "light",
        )
        item_mode = _mapped_code(
            values.get("Использование реквизита"),
            ITEM_MODE_NAMES,
            "none",
        )
        review_status, is_enabled, is_archived = _card_state_from_label(
            str(values.get("Состояние") or "Черновик").strip().casefold()
        )
        required_items = []
        for column in ("Реквизит 1", "Реквизит 2"):
            raw_item = str(values.get(column) or "").strip()
            if raw_item:
                required_items.append(item_codes.get(raw_item.casefold(), raw_item))
        risk_tags = _mapped_columns(values, "Тема риска", RISK_TAG_NAMES, 3)
        avoid_tags = _mapped_columns(values, "Исключить при границе", RISK_TAG_NAMES, 3)
        collections = ["restricted_content"] if extreme else [
            "kamasutra_inspired_poses" if category == "pose" else "base_tasks"
        ]
        rows.append(
            {
                "external_id": external_id,
                "level": str(level),
                "category": str(category),
                "intensity": str("hard" if extreme else intensity),
                "title": str(values.get("Название") or "").strip(),
                "text": text,
                "item_mode": str(item_mode),
                "required_items": ",".join(required_items),
                "timer_seconds": _integer_text(values.get("Таймер, сек.")),
                "risk_tags": ",".join(risk_tags),
                "avoid_if_tags": ",".join(avoid_tags),
                "collections": ",".join(collections),
                "review_status": review_status,
                "is_enabled": str(is_enabled),
                "_is_archived": str(is_archived),
                "aftercare_required": "1" if level == 4 or intensity == "hard" else "0",
                "pose_family": str(
                    _mapped_code(values.get("Семейство позы"), POSE_FAMILY_NAMES, "")
                ),
                "pose_difficulty": str(
                    _mapped_code(values.get("Сложность позы"), POSE_DIFFICULTY_NAMES, "")
                ),
                "space_required": str(_mapped_code(values.get("Место"), SPACE_NAMES, "")),
                "body_load": str(_mapped_code(values.get("Нагрузка"), BODY_LOAD_NAMES, "")),
            }
        )
    return rows


def _read_russian_item_sheet(sheet) -> list[dict[str, Any]]:
    rows = []
    for row_number, values in _sheet_dict_rows(sheet):
        name = str(values.get("Название") or "").strip()
        if not name:
            continue
        code = str(values.get("Код реквизита") or "").strip() or _generated_item_code(name)
        try:
            min_level = int(values.get("Минимальный уровень") or 1)
            max_level = int(values.get("Максимальный уровень") or 4)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Лист «Реквизит», строка {row_number}: уровни должны быть числами от 1 до 4."
            ) from exc
        if not 1 <= min_level <= max_level <= 4:
            raise ValueError(
                f"Лист «Реквизит», строка {row_number}: нужен диапазон уровней от 1 до 4."
            )
        categories = []
        for index in range(1, 4):
            value = values.get(f"Тип карточки {index}")
            if value in (None, ""):
                continue
            category = _mapped_code(value, CATEGORY_NAMES, "")
            if category and category not in categories:
                categories.append(str(category))
        if not categories:
            categories = ["task", "pose", "desire"]
        state = str(values.get("Состояние") or "Активен").strip().casefold()
        rows.append(
            {
                "code": code,
                "name": name,
                "min_level": min_level,
                "max_level": max_level,
                "categories": ",".join(categories),
                "usage_text": str(values.get("Инструкция для игроков") or "").strip(),
                "randomizable": 1 if _to_bool(values.get("Случайная подстановка")) else 0,
                "is_archived": 1 if state == "в архиве" else 0,
            }
        )
    return rows


def _mapped_code(value: object, labels: dict[object, str], default: object) -> object:
    if value in (None, ""):
        return default
    mapped = code_from_label(value, labels)
    return default if mapped is None else mapped


def _mapped_columns(
    values: dict[str, object],
    prefix: str,
    labels: dict[object, str],
    count: int,
) -> list[str]:
    result = []
    for index in range(1, count + 1):
        value = values.get(f"{prefix} {index}")
        if value in (None, ""):
            continue
        code = _mapped_code(value, labels, str(value).strip())
        if code and str(code) not in result:
            result.append(str(code))
    return result


def _card_state_from_label(value: str) -> tuple[str, int, int]:
    if value == "включена":
        return "approved", 1, 0
    if value == "на проверке":
        return "needs_review", 0, 0
    if value == "отключена":
        return "disabled", 0, 0
    if value == "в архиве":
        return "disabled", 0, 1
    return "draft", 0, 0


def _integer_text(value: object) -> str:
    if value in (None, ""):
        return ""
    return str(int(float(str(value).replace(",", "."))))


def _generated_item_code(name: str) -> str:
    digest = hashlib.sha1(name.strip().casefold().encode("utf-8")).hexdigest()[:12]
    return f"xlsx_item_{digest}"


def _to_bool(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "да"}

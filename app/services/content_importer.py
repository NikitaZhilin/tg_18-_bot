from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from app.domain import CATEGORIES, FORBIDDEN_RISK_TAGS, INTENSITIES, dump_json_list, parse_json_list
from app.labels import (
    BODY_LOAD_NAMES,
    CATEGORY_NAMES,
    INTENSITY_NAMES,
    ITEM_MODE_NAMES,
    POSE_DIFFICULTY_NAMES,
    POSE_FAMILY_NAMES,
    RISK_TAG_NAMES,
    SPACE_NAMES,
    code_from_label,
)
from app.storage import Database
from app.storage.repositories.admin_actions import AdminActionRepository
from app.storage.repositories.cards import CardRepository
from app.storage.repositories.items import ItemRepository


REQUIRED_COLUMNS = {"external_id", "level", "category", "text"}
RISK_KEYWORDS = {
    "breath_control": ("дыхани", "удуш", "перекры"),
    "choking": ("горло", "шею", "шея"),
    "unsafe_wax": ("воск", "свеч"),
    "no_quick_release_restraint": ("свяж", "связ", "верев"),
    "unbounded_humiliation": ("униж",),
    "intoxication": ("алког", "опьян"),
    "recording_without_consent": ("фото", "видео", "запиш"),
    "public_nonconsenting_people": ("публич", "посторон"),
    "injury": ("боль", "удар", "след", "синяк"),
}
ITEM_ALIASES = {
    "лед": "ice",
    "масло": "oil",
    "веревка": "rope",
    "верёвка": "rope",
    "повязка": "blindfold",
    "вибратор": "vibrator",
    "наушники": "headphones",
    "свеча": "candle",
    "свечи": "candle",
    "зажимы": "clamps",
    "еда": "food",
}


@dataclass
class ImportWarning:
    row_number: int
    external_id: str | None
    message: str


@dataclass
class ImportReport:
    source_file: str
    content_version: str
    added_or_updated: int = 0
    items_added_or_updated: int = 0
    disabled_cards: int = 0
    needs_review: int = 0
    warnings: list[ImportWarning] = field(default_factory=list)
    dry_run: bool = False

    @property
    def warnings_count(self) -> int:
        return len(self.warnings)


class ContentImporter:
    def __init__(self, db: Database):
        self.db = db
        self.cards = CardRepository(db)
        self.items = ItemRepository(db)
        self.admin_actions = AdminActionRepository(db)

    def import_file(
        self,
        path: str | Path,
        *,
        content_version: str = "manual",
        dry_run: bool = False,
        admin_user_id: int | None = None,
        skip_existing: bool = False,
        preserve_admin_changes: bool = False,
        skip_imported_version: bool = False,
    ) -> ImportReport:
        path = Path(path)
        report = ImportReport(str(path), content_version, dry_run=dry_run)
        if skip_imported_version and self.db.fetchone(
            "SELECT 1 FROM content_imports WHERE content_version = ? LIMIT 1",
            (content_version,),
        ):
            return report

        item_rows: list[dict[str, Any]] = []
        if path.suffix.lower() == ".csv":
            rows = self._read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            rows, item_rows = self._read_xlsx(path)
        elif path.suffix.lower() == ".docx":
            rows = self._read_docx(path)
        else:
            raise ValueError(f"unsupported content file: {path.suffix}")

        prepared_rows: list[tuple[dict[str, Any], list[str], list[str]]] = []
        trusted_seed = path.parent.resolve() == Path("content").resolve()
        existing_external_ids = {
            str(row["external_id"])
            for row in self.db.fetchall(
                "SELECT external_id FROM cards WHERE external_id IS NOT NULL"
            )
        }
        admin_modified_external_ids = {
            str(row["external_id"])
            for row in self.db.fetchall(
                """
                SELECT DISTINCT c.external_id
                FROM cards c
                JOIN card_versions cv ON cv.card_id = c.id
                WHERE c.external_id IS NOT NULL
                """
            )
        } if preserve_admin_changes else set()
        known_item_codes = {
            str(row["code"])
            for row in self.db.fetchall(
                "SELECT code FROM items WHERE deleted_at IS NULL"
            )
        }
        known_item_codes.update(
            str(item["code"])
            for item in item_rows
            if str(item.get("code") or "").strip()
        )

        for index, row in enumerate(rows, start=2):
            external_id = str(row.get("external_id") or "").strip()
            if skip_existing and external_id in existing_external_ids:
                continue
            if external_id in admin_modified_external_ids:
                continue
            if trusted_seed:
                row["_trusted_seed"] = "1"
            try:
                card_data, items, collections = self._prepare_row(row)
            except ValueError as exc:
                report.warnings.append(ImportWarning(index, row.get("external_id"), str(exc)))
                continue
            unknown_items = sorted(set(items) - known_item_codes)
            if unknown_items:
                report.warnings.append(
                    ImportWarning(
                        index,
                        card_data["external_id"],
                        "Неизвестный реквизит: " + ", ".join(unknown_items),
                    )
                )
                continue
            if card_data["review_status"] == "disabled":
                report.disabled_cards += 1
            if card_data["review_status"] == "needs_review":
                report.needs_review += 1
            prepared_rows.append((card_data, items, collections))

        report.added_or_updated = len(prepared_rows)
        report.items_added_or_updated = len(item_rows)
        if dry_run:
            return report

        with self.db.transaction() as conn:
            for item_data in item_rows:
                self.items.upsert_from_import(item_data, conn)
            for card_data, items, collections in prepared_rows:
                card_id = self.cards.upsert(card_data, items, collections, conn)
                self.cards.set_archived(card_id, bool(card_data.get("_is_archived")), conn)
            conn.execute(
                """
                INSERT INTO content_imports (
                    source_file, content_version, imported_cards, disabled_cards, warnings_count, report_json
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    str(path),
                    content_version,
                    report.added_or_updated,
                    report.disabled_cards,
                    report.warnings_count,
                    "{}",
                ),
            )
        if admin_user_id:
            self.admin_actions.record(
                admin_user_id,
                "import_cards",
                "content_file",
                str(path),
                {
                    "content_version": content_version,
                    "imported_cards": report.added_or_updated,
                    "disabled_cards": report.disabled_cards,
                    "warnings_count": report.warnings_count,
                    "items_added_or_updated": report.items_added_or_updated,
                },
            )
        return report

    @staticmethod
    def _read_csv(path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            if not reader.fieldnames or not REQUIRED_COLUMNS.issubset(set(reader.fieldnames)):
                missing = ", ".join(sorted(REQUIRED_COLUMNS - set(reader.fieldnames or [])))
                raise ValueError(f"missing required columns: {missing}")
            return [dict(row) for row in reader]

    @staticmethod
    def _read_xlsx(path: Path) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if wb.active and _english_headers(wb.active):
                rows = list(wb.active.iter_rows(values_only=True))
                headers = [str(cell or "").strip() for cell in rows[0]]
                return (
                    [
                        {
                            headers[i]: "" if i >= len(raw) or raw[i] is None else str(raw[i])
                            for i in range(len(headers))
                        }
                        for raw in rows[1:]
                    ],
                    [],
                )

            item_rows = _read_russian_item_sheet(wb["Реквизит"]) if "Реквизит" in wb.sheetnames else []
            item_codes = {
                str(item["name"]).strip().casefold(): str(item["code"])
                for item in item_rows
            }
            result: list[dict[str, str]] = []
            sheet_levels = {
                "Флирт": 1,
                "Разогрев": 2,
                "Секс": 3,
                "BDSM": 4,
                "Экстрим": 4,
                "Закрытые темы": 4,
            }
            for sheet_name, level in sheet_levels.items():
                if sheet_name not in wb.sheetnames:
                    continue
                result.extend(
                    _read_russian_card_sheet(
                        wb[sheet_name],
                        level=level,
                        extreme=sheet_name in {"Экстрим", "Закрытые темы"},
                        item_codes=item_codes,
                    )
                )
            if not result:
                raise ValueError(
                    "В XLSX не найдены листы «Флирт», «Разогрев», «Секс», «BDSM» или «Экстрим»."
                )
            return result, item_rows
        finally:
            wb.close()

    @staticmethod
    def _read_docx(path: Path) -> list[dict[str, str]]:
        from docx import Document

        doc = Document(path)
        rows: list[dict[str, str]] = []
        level = 1
        counter = 1
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            lower = text.lower()
            if "уровень 1" in lower:
                level = 1
                continue
            if "уровень 2" in lower:
                level = 2
                continue
            if "уровень 3" in lower:
                level = 3
                continue
            if "уровень 4" in lower:
                level = 4
                continue
            if len(text) < 12:
                continue
            required_items = _extract_items(text)
            timer_seconds = _extract_timer_seconds(text)
            risk_tags = _detect_risk_tags(text)
            rows.append(
                {
                    "external_id": f"docx_l{level}_{counter:03d}",
                    "level": str(level),
                    "category": "task",
                    "intensity": "light" if level < 3 else "medium",
                    "title": "",
                    "text": _strip_docx_metadata(text),
                    "required_items": ",".join(required_items),
                    "timer_seconds": str(timer_seconds or ""),
                    "risk_tags": ",".join(risk_tags),
                    "collection": "base_tasks",
                    "review_status": "needs_review",
                    "is_enabled": "0",
                }
            )
            counter += 1
        return rows

    @staticmethod
    def _prepare_row(row: dict[str, Any]) -> tuple[dict[str, Any], list[str], list[str]]:
        external_id = str(row.get("external_id") or "").strip()
        if not external_id:
            raise ValueError("external_id is required")
        text = str(row.get("text") or "").strip()
        if not text:
            raise ValueError("text is required")
        level = int(str(row.get("level") or "1").strip())
        if level < 1 or level > 4:
            raise ValueError("level must be 1..4")
        category = str(row.get("category") or "task").strip()
        if category not in CATEGORIES:
            raise ValueError(f"unknown category: {category}")
        intensity = str(row.get("intensity") or "light").strip()
        if intensity not in INTENSITIES:
            raise ValueError(f"unknown intensity: {intensity}")
        item_mode = str(row.get("item_mode") or "none").strip()
        if item_mode not in ITEM_MODE_NAMES:
            raise ValueError(f"unknown item mode: {item_mode}")

        risk_tags = set(parse_json_list(row.get("risk_tags")))
        avoid_if_tags = parse_json_list(row.get("avoid_if_tags"))
        forbidden = risk_tags.intersection(FORBIDDEN_RISK_TAGS)
        review_status = str(row.get("review_status") or "draft").strip()
        is_enabled = _to_bool(row.get("is_enabled"), default=False)
        requires_both_opt_in = _to_bool(row.get("requires_both_opt_in"), default=False)
        requires_safeword_check = _to_bool(row.get("requires_safeword_check"), default=False)
        aftercare_required = _to_bool(row.get("aftercare_required"), default=False)

        if level == 4 or intensity == "hard":
            requires_both_opt_in = True
            requires_safeword_check = True
            aftercare_required = True
            if _to_bool(row.get("_trusted_seed")) and review_status == "needs_review":
                review_status = "approved"
                is_enabled = True
        if forbidden:
            is_enabled = False
            review_status = "disabled"
        elif review_status != "approved":
            is_enabled = False

        if category == "pose":
            for field_name in ("pose_family", "pose_difficulty", "space_required", "body_load"):
                if not str(row.get(field_name) or "").strip():
                    raise ValueError(f"{field_name} is required for pose")

        timer_seconds = str(row.get("timer_seconds") or "").strip()
        card_data = {
            "external_id": external_id,
            "level": level,
            "category": category,
            "intensity": intensity,
            "title": _empty_to_none(row.get("title")),
            "text": text,
            "media_id": _empty_to_none(row.get("media_id")),
            "media_file": _empty_to_none(row.get("media_file")),
            "media_type": _empty_to_none(row.get("media_type")),
            "pose_family": _empty_to_none(row.get("pose_family")),
            "pose_difficulty": _empty_to_none(row.get("pose_difficulty")),
            "space_required": _empty_to_none(row.get("space_required")),
            "body_load": _empty_to_none(row.get("body_load")),
            "avoid_if_tags": dump_json_list(avoid_if_tags),
            "source_note": _empty_to_none(row.get("source_note")),
            "timer_seconds": int(timer_seconds) if timer_seconds else None,
            "safety_level": str(row.get("safety_level") or "normal").strip(),
            "risk_tags": dump_json_list(risk_tags),
            "requires_both_opt_in": 1 if requires_both_opt_in else 0,
            "requires_safeword_check": 1 if requires_safeword_check else 0,
            "aftercare_required": 1 if aftercare_required else 0,
            "item_mode": item_mode,
            "is_enabled": 1 if is_enabled else 0,
            "review_status": review_status,
            "notes": _empty_to_none(row.get("notes")),
            "_is_archived": 1 if _to_bool(row.get("_is_archived")) else 0,
        }
        items = parse_json_list(row.get("required_items"))
        if items and item_mode != "required":
            raise ValueError(
                "При явно указанном реквизите выберите режим «Обязательно подобрать»"
            )
        collections = parse_json_list(row.get("collections") or row.get("collection"))
        if not collections:
            collections = ["base_tasks"]
        return card_data, items, collections


def _empty_to_none(value: object) -> str | None:
    raw = str(value or "").strip()
    return raw or None


def _to_bool(value: object, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "да"}


def _extract_items(text: str) -> list[str]:
    match = re.search(r"\[?\s*Реквизит\s*:\s*([^\]\n]+)\]?", text, flags=re.IGNORECASE)
    if not match:
        return []
    raw_value = re.split(r"\bТаймер\s*:", match.group(1), flags=re.IGNORECASE)[0]
    raw_items = parse_json_list(raw_value)
    result = []
    for raw in raw_items:
        key = raw.lower().strip()
        result.append(ITEM_ALIASES.get(key, key))
    return result


def _extract_timer_seconds(text: str) -> int | None:
    match = re.search(r"\[?\s*Таймер\s*:\s*(\d+)\s*(сек|секунд|мин|минут)?", text, flags=re.IGNORECASE)
    if not match:
        return None
    value = int(match.group(1))
    unit = (match.group(2) or "сек").lower()
    return value * 60 if unit.startswith("мин") else value


def _detect_risk_tags(text: str) -> list[str]:
    lower = text.lower()
    detected = []
    for tag, keywords in RISK_KEYWORDS.items():
        if any(keyword in lower for keyword in keywords):
            detected.append(tag)
    return detected


def _strip_docx_metadata(text: str) -> str:
    text = re.sub(r"\[?\s*Реквизит\s*:\s*[^\]\n]+\]?", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\[?\s*Таймер\s*:\s*[^\]\n]+\]?", "", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def _english_headers(sheet) -> bool:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {str(value or "").strip() for value in first_row}
    return REQUIRED_COLUMNS.issubset(headers)


def _sheet_dict_rows(sheet) -> list[tuple[int, dict[str, object]]]:
    raw_rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(raw_rows, ())]
    result = []
    for row_number, raw in enumerate(raw_rows, start=2):
        values = {
            header: raw[index] if index < len(raw) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in values.values()):
            result.append((row_number, values))
    return result


def _read_russian_card_sheet(
    sheet,
    *,
    level: int,
    extreme: bool,
    item_codes: dict[str, str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_number, values in _sheet_dict_rows(sheet):
        text = str(values.get("Текст карточки") or "").strip()
        if not text:
            continue
        external_id = str(values.get("Код карточки") or "").strip()
        if not external_id:
            digest = hashlib.sha1(
                f"{sheet.title}|{row_number}|{text}".encode("utf-8")
            ).hexdigest()[:12]
            external_id = f"xlsx_{digest}"
        category = _mapped_code(values.get("Тип карточки"), CATEGORY_NAMES, "task")
        intensity = _mapped_code(values.get("Интенсивность"), INTENSITY_NAMES, "hard" if extreme else "light")
        item_mode = _mapped_code(values.get("Использование реквизита"), ITEM_MODE_NAMES, "none")
        state = str(values.get("Состояние") or "Черновик").strip().casefold()
        review_status, is_enabled, is_archived = _card_state_from_label(state)

        required_items = []
        for column in ("Реквизит 1", "Реквизит 2"):
            raw_item = str(values.get(column) or "").strip()
            if not raw_item:
                continue
            required_items.append(item_codes.get(raw_item.casefold(), raw_item))

        risk_tags = _mapped_columns(values, "Тема риска", RISK_TAG_NAMES, 3)
        avoid_tags = _mapped_columns(values, "Исключить при границе", RISK_TAG_NAMES, 3)
        collections = ["restricted_content"] if extreme else [
            "kamasutra_inspired_poses" if category == "pose" else "base_tasks"
        ]
        rows.append(
            {
                "external_id": external_id,
                "level": str(level),
                "category": str(category),
                "intensity": str("hard" if extreme else intensity),
                "title": str(values.get("Название") or "").strip(),
                "text": text,
                "item_mode": str(item_mode),
                "required_items": ",".join(required_items),
                "timer_seconds": _integer_text(values.get("Таймер, сек.")),
                "risk_tags": ",".join(risk_tags),
                "avoid_if_tags": ",".join(avoid_tags),
                "collections": ",".join(collections),
                "review_status": review_status,
                "is_enabled": str(is_enabled),
                "_is_archived": str(is_archived),
                "requires_both_opt_in": "1" if level == 4 or intensity == "hard" else "0",
                "requires_safeword_check": "1" if level == 4 or intensity == "hard" else "0",
                "aftercare_required": "1" if level == 4 or intensity == "hard" else "0",
                "safety_level": "hard" if level == 4 or intensity == "hard" else "normal",
                "pose_family": str(
                    _mapped_code(values.get("Семейство позы"), POSE_FAMILY_NAMES, "")
                ),
                "pose_difficulty": str(
                    _mapped_code(values.get("Сложность позы"), POSE_DIFFICULTY_NAMES, "")
                ),
                "space_required": str(_mapped_code(values.get("Место"), SPACE_NAMES, "")),
                "body_load": str(_mapped_code(values.get("Нагрузка"), BODY_LOAD_NAMES, "")),
            }
        )
    return rows


def _read_russian_item_sheet(sheet) -> list[dict[str, Any]]:
    rows = []
    for row_number, values in _sheet_dict_rows(sheet):
        name = str(values.get("Название") or "").strip()
        if not name:
            continue
        code = str(values.get("Код реквизита") or "").strip() or _generated_item_code(name)
        try:
            min_level = int(values.get("Минимальный уровень") or 1)
            max_level = int(values.get("Максимальный уровень") or 4)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Лист «Реквизит», строка {row_number}: уровни должны быть числами от 1 до 4."
            ) from exc
        if not 1 <= min_level <= max_level <= 4:
            raise ValueError(
                f"Лист «Реквизит», строка {row_number}: нужен диапазон уровней от 1 до 4."
            )
        categories = []
        for index in range(1, 4):
            value = values.get(f"Тип карточки {index}")
            if value in (None, ""):
                continue
            category = _mapped_code(value, CATEGORY_NAMES, "")
            if category and category not in categories:
                categories.append(str(category))
        if not categories:
            categories = ["task", "pose", "desire"]
        state = str(values.get("Состояние") or "Активен").strip().casefold()
        rows.append(
            {
                "code": code,
                "name": name,
                "min_level": min_level,
                "max_level": max_level,
                "categories": ",".join(categories),
                "usage_text": str(values.get("Инструкция для игроков") or "").strip(),
                "randomizable": 1 if _to_bool(values.get("Случайная подстановка")) else 0,
                "is_archived": 1 if state == "в архиве" else 0,
            }
        )
    return rows


def _mapped_code(value: object, labels: dict[object, str], default: object) -> object:
    if value in (None, ""):
        return default
    mapped = code_from_label(value, labels)
    return default if mapped is None else mapped


def _mapped_columns(
    values: dict[str, object],
    prefix: str,
    labels: dict[object, str],
    count: int,
) -> list[str]:
    result = []
    for index in range(1, count + 1):
        value = values.get(f"{prefix} {index}")
        if value in (None, ""):
            continue
        code = _mapped_code(value, labels, str(value).strip())
        if code and str(code) not in result:
            result.append(str(code))
    return result


def _card_state_from_label(value: str) -> tuple[str, int, int]:
    if value == "включена":
        return "approved", 1, 0
    if value == "на проверке":
        return "needs_review", 0, 0
    if value == "отключена":
        return "disabled", 0, 0
    if value == "в архиве":
        return "disabled", 0, 1
    return "draft", 0, 0


def _integer_text(value: object) -> str:
    if value in (None, ""):
        return ""
    return str(int(float(str(value).replace(",", "."))))


def _generated_item_code(name: str) -> str:
    digest = hashlib.sha1(name.strip().casefold().encode("utf-8")).hexdigest()[:12]
    return f"xlsx_item_{digest}"

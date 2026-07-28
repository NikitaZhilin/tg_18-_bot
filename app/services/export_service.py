from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.domain import parse_json_list
from app.labels import (
    BODY_LOAD_NAMES,
    CATEGORY_NAMES,
    INTENSITY_NAMES,
    ITEM_MODE_NAMES,
    POSE_DIFFICULTY_NAMES,
    POSE_FAMILY_NAMES,
    RISK_TAG_NAMES,
    SPACE_NAMES,
)


CARD_COLUMNS = [
    "Код карточки",
    "Тип карточки",
    "Интенсивность",
    "Название",
    "Текст карточки",
    "Использование реквизита",
    "Реквизит 1",
    "Реквизит 2",
    "Таймер, сек.",
    "Тема риска 1",
    "Тема риска 2",
    "Тема риска 3",
    "Исключить при границе 1",
    "Исключить при границе 2",
    "Исключить при границе 3",
    "Состояние",
    "Семейство позы",
    "Сложность позы",
    "Место",
    "Нагрузка",
]

ITEM_COLUMNS = [
    "Код реквизита",
    "Название",
    "Минимальный уровень",
    "Максимальный уровень",
    "Тип карточки 1",
    "Тип карточки 2",
    "Тип карточки 3",
    "Инструкция для игроков",
    "Случайная подстановка",
    "Состояние",
]

LEVEL_SHEETS = {
    1: "Флирт",
    2: "Разогрев",
    3: "Секс",
    4: "BDSM",
}

EXPORT_COLUMNS = CARD_COLUMNS

HEADER_FILL = PatternFill("solid", fgColor="165B65")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="E7F1F2")
ARCHIVE_FILL = PatternFill("solid", fgColor="E5E7EB")
THIN_BORDER = Border(bottom=Side(style="thin", color="B8C6C8"))


def save_cards_xlsx(
    rows: Sequence[Mapping[str, object]],
    path: str | Path,
    items: Sequence[Mapping[str, object]] = (),
) -> None:
    path = Path(path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    items_by_code = {
        str(item["code"]): str(item["name"])
        for item in items
    }

    card_sheets = {
        name: workbook.create_sheet(name)
        for name in [*LEVEL_SHEETS.values(), "Экстрим"]
    }
    item_sheet = workbook.create_sheet("Реквизит")
    reference_sheet = workbook.create_sheet("Справочники")

    for sheet in card_sheets.values():
        sheet.append(CARD_COLUMNS)

    for row in rows:
        collections = set(parse_json_list(_row_value(row, "collections")))
        is_extreme = "restricted_content" in collections
        target = card_sheets["Экстрим"] if is_extreme else card_sheets[LEVEL_SHEETS[int(row["level"])]]
        target.append(_export_card_row(row, items_by_code))

    item_sheet.append(ITEM_COLUMNS)
    for item in items:
        item_sheet.append(_export_item_row(item))

    _write_reference_sheet(reference_sheet)
    _create_defined_names(workbook, reference_sheet, item_sheet)

    for index, (name, sheet) in enumerate(card_sheets.items(), start=1):
        _style_card_sheet(sheet, table_name=f"Cards{index}")
        _add_card_validations(sheet)
    _style_item_sheet(item_sheet)
    _add_item_validations(item_sheet)
    _style_reference_sheet(reference_sheet)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def _export_card_row(
    row: Mapping[str, object],
    items_by_code: Mapping[str, str],
) -> list[object]:
    required_items = [
        items_by_code.get(code, code)
        for code in parse_json_list(_row_value(row, "required_items"))
    ]
    risk_tags = [
        RISK_TAG_NAMES.get(code, code)
        for code in parse_json_list(_row_value(row, "risk_tags"))
    ]
    avoid_tags = [
        RISK_TAG_NAMES.get(code, code)
        for code in parse_json_list(_row_value(row, "avoid_if_tags"))
    ]
    return [
        _row_value(row, "external_id"),
        CATEGORY_NAMES.get(str(_row_value(row, "category")), _row_value(row, "category")),
        INTENSITY_NAMES.get(str(_row_value(row, "intensity")), str(_row_value(row, "intensity"))).capitalize(),
        _row_value(row, "title"),
        _row_value(row, "text"),
        ITEM_MODE_NAMES.get(str(_row_value(row, "item_mode") or "none"), "Не использовать"),
        required_items[0] if len(required_items) > 0 else "",
        required_items[1] if len(required_items) > 1 else "",
        _row_value(row, "timer_seconds"),
        risk_tags[0] if len(risk_tags) > 0 else "",
        risk_tags[1] if len(risk_tags) > 1 else "",
        risk_tags[2] if len(risk_tags) > 2 else "",
        avoid_tags[0] if len(avoid_tags) > 0 else "",
        avoid_tags[1] if len(avoid_tags) > 1 else "",
        avoid_tags[2] if len(avoid_tags) > 2 else "",
        _card_state(row),
        POSE_FAMILY_NAMES.get(str(_row_value(row, "pose_family")), _row_value(row, "pose_family")),
        POSE_DIFFICULTY_NAMES.get(
            str(_row_value(row, "pose_difficulty")),
            _row_value(row, "pose_difficulty"),
        ),
        SPACE_NAMES.get(str(_row_value(row, "space_required")), _row_value(row, "space_required")),
        BODY_LOAD_NAMES.get(str(_row_value(row, "body_load")), _row_value(row, "body_load")),
    ]


def _export_item_row(item: Mapping[str, object]) -> list[object]:
    categories = parse_json_list(_row_value(item, "categories"))
    return [
        _row_value(item, "code"),
        _row_value(item, "name"),
        _row_value(item, "min_level"),
        _row_value(item, "max_level"),
        CATEGORY_NAMES.get(categories[0], categories[0]) if len(categories) > 0 else "",
        CATEGORY_NAMES.get(categories[1], categories[1]) if len(categories) > 1 else "",
        CATEGORY_NAMES.get(categories[2], categories[2]) if len(categories) > 2 else "",
        _row_value(item, "usage_text"),
        "Да" if int(_row_value(item, "randomizable") or 0) else "Нет",
        "В архиве" if int(_row_value(item, "is_archived") or 0) else "Активен",
    ]


def _card_state(row: Mapping[str, object]) -> str:
    if int(_row_value(row, "is_archived") or 0):
        return "В архиве"
    status = str(_row_value(row, "review_status") or "draft")
    if status == "draft":
        return "Черновик"
    if status == "needs_review":
        return "На проверке"
    if status == "approved" and int(_row_value(row, "is_enabled") or 0):
        return "Включена"
    return "Отключена"


def _style_card_sheet(sheet, *, table_name: str) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(CARD_COLUMNS))}{max(sheet.max_row, 2)}"
    _style_header(sheet, len(CARD_COLUMNS))
    widths = [24, 18, 16, 24, 72, 24, 22, 22, 14, 24, 24, 24, 26, 26, 26, 18, 24, 18, 18, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    sheet.row_dimensions[1].height = 34
    if sheet.max_row >= 2:
        table = Table(
            displayName=table_name,
            ref=f"A1:{get_column_letter(len(CARD_COLUMNS))}{sheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
        )
        sheet.add_table(table)
    archive_rule = FormulaRule(formula=["$P2=\"В архиве\""], fill=ARCHIVE_FILL)
    sheet.conditional_formatting.add(f"A2:T{max(sheet.max_row, 1000)}", archive_rule)


def _style_item_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    _style_header(sheet, len(ITEM_COLUMNS))
    widths = [28, 24, 18, 18, 20, 20, 20, 72, 24, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    if sheet.max_row >= 2:
        table = Table(displayName="ItemsCatalog", ref=f"A1:J{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        sheet.add_table(table)


def _style_header(sheet, column_count: int) -> None:
    for cell in sheet[1][:column_count]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet["A1"].comment = Comment(
        "Код нужен для обновления существующей записи. Оставьте его пустым, чтобы создать новую.",
        "Codex",
    )


def _write_reference_sheet(sheet) -> None:
    columns = {
        "A": ("Типы карточек", list(CATEGORY_NAMES.values())),
        "B": ("Интенсивность", [value.capitalize() for value in INTENSITY_NAMES.values()]),
        "C": ("Использование реквизита", list(ITEM_MODE_NAMES.values())),
        "D": ("Состояние карточки", ["Черновик", "На проверке", "Включена", "Отключена", "В архиве"]),
        "E": ("Темы риска", list(RISK_TAG_NAMES.values())),
        "F": ("Сложность позы", list(POSE_DIFFICULTY_NAMES.values())),
        "G": ("Место", list(SPACE_NAMES.values())),
        "H": ("Нагрузка", list(BODY_LOAD_NAMES.values())),
        "I": ("Семейство позы", list(POSE_FAMILY_NAMES.values())),
        "J": ("Да / Нет", ["Да", "Нет"]),
        "K": ("Уровни", [1, 2, 3, 4]),
        "L": ("Состояние реквизита", ["Активен", "В архиве"]),
    }
    for column, (title, values) in columns.items():
        sheet[f"{column}1"] = title
        for row_index, value in enumerate(values, start=2):
            sheet[f"{column}{row_index}"] = value
    sheet["N1"] = "Тема риска"
    sheet["O1"] = "Что означает"
    for row_index, (code, label) in enumerate(RISK_TAG_NAMES.items(), start=2):
        sheet[f"N{row_index}"] = label
        sheet[f"O{row_index}"] = _risk_explanation(code)


def _style_reference_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, 13):
        sheet.column_dimensions[get_column_letter(column)].width = 26
    sheet.column_dimensions["N"].width = 30
    sheet.column_dimensions["O"].width = 72
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _create_defined_names(workbook, reference_sheet, item_sheet) -> None:
    definitions = {
        "CardTypes": ("A", len(CATEGORY_NAMES)),
        "Intensities": ("B", len(INTENSITY_NAMES)),
        "ItemModes": ("C", len(ITEM_MODE_NAMES)),
        "CardStates": ("D", 5),
        "RiskTags": ("E", len(RISK_TAG_NAMES)),
        "PoseDifficulties": ("F", len(POSE_DIFFICULTY_NAMES)),
        "Spaces": ("G", len(SPACE_NAMES)),
        "BodyLoads": ("H", len(BODY_LOAD_NAMES)),
        "PoseFamilies": ("I", len(POSE_FAMILY_NAMES)),
        "YesNo": ("J", 2),
        "Levels": ("K", 4),
        "ItemStates": ("L", 2),
    }
    for name, (column, count) in definitions.items():
        workbook.defined_names.add(
            DefinedName(
                name,
                attr_text=f"{quote_sheetname(reference_sheet.title)}!${column}$2:${column}${count + 1}",
            )
        )
    workbook.defined_names.add(
        DefinedName(
            "ItemNames",
            attr_text=f"{quote_sheetname(item_sheet.title)}!$B$2:$B${max(item_sheet.max_row, 2)}",
        )
    )


def _add_card_validations(sheet) -> None:
    validations = [
        ("B2:B1000", "CardTypes"),
        ("C2:C1000", "Intensities"),
        ("F2:F1000", "ItemModes"),
        ("G2:H1000", "ItemNames"),
        ("J2:O1000", "RiskTags"),
        ("P2:P1000", "CardStates"),
        ("Q2:Q1000", "PoseFamilies"),
        ("R2:R1000", "PoseDifficulties"),
        ("S2:S1000", "Spaces"),
        ("T2:T1000", "BodyLoads"),
    ]
    for cell_range, defined_name in validations:
        validation = DataValidation(type="list", formula1=f"={defined_name}", allow_blank=True)
        validation.error = "Выберите значение из списка."
        validation.errorTitle = "Недопустимое значение"
        validation.prompt = "Используйте раскрывающийся список."
        validation.promptTitle = "Выбор значения"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(cell_range)
    timer_validation = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    timer_validation.error = "Введите целое количество секунд больше нуля."
    sheet.add_data_validation(timer_validation)
    timer_validation.add("I2:I1000")


def _add_item_validations(sheet) -> None:
    for cell_range, defined_name in (
        ("C2:D1000", "Levels"),
        ("E2:G1000", "CardTypes"),
        ("I2:I1000", "YesNo"),
        ("J2:J1000", "ItemStates"),
    ):
        validation = DataValidation(type="list", formula1=f"={defined_name}", allow_blank=True)
        validation.error = "Выберите значение из списка."
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(cell_range)


def _risk_explanation(code: str) -> str:
    explanations = {
        "power_exchange": "Один партнер временно ведет сцену, второй следует заранее согласованным правилам.",
        "roleplay": "Действия происходят в выбранных ролях с ясным началом и завершением.",
        "command_language": "В карточке используется строгий или командный тон.",
        "denial_play": "Один партнер временно ограничивает свою инициативу в согласованных пределах.",
        "consent_check": "Во время карточки нужно несколько раз прямо проверить комфорт и согласие.",
        "pose_control": "Один партнер задает или ограничивает положение тела другого.",
        "sensory_deprivation": "Повязка, наушники или иное временное ограничение органов чувств.",
        "food": "В задании может использоваться еда или напитки.",
        "toys": "В задании могут использоваться интимные игрушки.",
        "fisting": "Практика для опытной пары, требующая отдельной подготовки и немедленной остановки при боли.",
        "advanced_insertion": "Сложная практика проникновения, требующая опыта и отдельного согласия.",
        "urethral_play": "Уретральная практика с повышенными требованиями к стерильности.",
        "sterile_equipment": "Нужен новый или стерильный специализированный реквизит.",
        "advanced_practice": "Карточка предназначена только для пары с предыдущим опытом.",
        "aftercare": "После задания нужно полностью остановиться и спокойно обсудить состояние.",
        "injury": "Карточка исключается при запрете боли, риска травмы или следов.",
        "medical_condition": "Карточка исключается при соответствующих ограничениях по здоровью.",
        "urinary_symptoms": "Карточка исключается при боли, воспалении или других симптомах мочевыводящих путей.",
        "no_quick_release_restraint": "В карточке есть фиксация или ограничение движений.",
        "unbounded_humiliation": "В карточке может быть унизительная лексика или роль.",
        "unsafe_wax": "В карточке есть горячий воск или открытое пламя.",
    }
    return explanations.get(code, "")


def _row_value(row: Mapping[str, object], key: str) -> object:
    try:
        return row[key]
    except (KeyError, IndexError):
        return ""
